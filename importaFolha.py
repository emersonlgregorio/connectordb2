import pandas as pd
from openpyxl import load_workbook
import base64
import os


def decodeArquivo(file):
    arquivo = file.split(';')[2].split(',')[1]
    arquivo = base64.b64decode(arquivo)
    output_file = open('arquivo_processado.xlsx', 'wb')
    output_file.write(arquivo)
    output_file.close()


def preparaArquivo(file):
    print('Iniciando importanção do arquivo. Processo: Importacao Folha de Pagamento')

    if os.path.exists('arquivo_processado.xlsx'):
        os.remove('arquivo_processado.xlsx')

    decodeArquivo(file)

    arquivo = 'arquivo_processado.xlsx'
    wb = load_workbook(arquivo)
    ws = wb.active

    # Cria lista de ajustes
    listaAjustes = []
    listaExclui = []
    for contador, i in enumerate(ws.iter_rows(min_row=0,
                                              # max_row=100,
                                              values_only=True)):

        if i[5] == 'Pág.:':
            pagina = i[6]
            empresa = i[0]
        if i[0] == 'Filial:':
            filial = i[1]
            ccusto = i[3]
        if i[0] == 'Conta':
            linhaInicial = contador + 1
        if i[0] == 'Débitos:':
            linhaFinal = contador
            listaAjustes.append(
                {
                    "pagina": pagina,
                    "empresa": empresa,
                    "filial": filial,
                    "ccusto": ccusto,
                    "linhaInicial": linhaInicial,
                    "linhaFinal": linhaFinal
                }
            )
    # print(listaAjustes)

    # Faz ajustes no arquivo e grava com novo nome
    for l in listaAjustes:
        l_ini = l['linhaInicial']
        l_fin = l['linhaFinal']
        filial = l['filial']
        empresa = l['empresa']
        # print(l['filial'], l['empresa'], filial)
        ccusto = l['ccusto']
        intervalo = range(l_ini, l_fin)
        for contador, i in enumerate(ws.iter_rows(min_row=0,

                                                  # max_row=100,
                                                  values_only=True)):
            # print(intervalo)
            c = contador + 1
            if contador in intervalo:
                # print(contador, i, l_ini, l_fin)
                ws[f'K{c}'] = empresa
                ws[f'L{c}'] = filial
                ws[f'M{c}'] = ccusto
    wb.save(arquivo)
    wb.close()

    # Marca Linhas para exclusão
    wb2 = load_workbook(arquivo)
    ws2 = wb2.active
    ws2['N1'] = 'Excluir'
    wb2.save(arquivo)
    for contador, i in enumerate(ws2.iter_rows(min_row=2,
                                               # max_row=100,
                                               values_only=True)):
        # print(contador, i[11])
        c = contador + 2
        if (not i[11]) or (i[11] == None) or (i[11] == '') or (i[0] == 'Conta'):
            ws2[f'N{c}'] = 'Excluir'
            # ws.delete_rows(contador)
    wb2.save(arquivo)
    wb2.close()

    # Exclui linhas do arquivo
    df = pd.read_excel(arquivo)
    filtro = df['Excluir'] != 'Excluir'
    folha_contabil = df[filtro]
    folha_contabil.to_excel(arquivo)

    # transformar em json
    wb3 = load_workbook(arquivo)
    ws3 = wb3.active
    folhaContabil = []
    for contador, i in enumerate(ws3.iter_rows(min_row=2,
                                               # max_row=100,
                                               values_only=True)):
        folhaContabil.append(
            {
                "conta1": str(i[2]),
                "conta2": str(i[3]),
                "nomeEvento": i[6],
                "valor": i[7],
                "tipoLancamento": i[8],
                "codigoEvento": i[10],
                "empresa": i[11],
                "filial": i[12],
                "ccusto": i[13]

            }
        )

    print('Importanção do arquivo finalizada. Processo: Importacao Folha de Pagamento')
    wb3.close()
    return folhaContabil

# if __name__ == '__main__':
#     teste = preparaArquivo('contabilizacao_de_ janeiro.xlsx')
#     print(teste)